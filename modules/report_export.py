# -*- coding: utf-8 -*-
"""按前端日期范围导出黄阁水厂碳排放 Excel 报告。"""

from datetime import date
from io import BytesIO
from pathlib import Path
from urllib.parse import quote

import pandas as pd
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel


router = APIRouter()

BASE_DIR = Path(__file__).resolve().parent.parent
REALTIME_DIR = BASE_DIR / "data" / "real-time output"
HISTORY_DIR = BASE_DIR / "data" / "history"
HISTORY_SCOPE_FILE = HISTORY_DIR / "scope123_daily.csv"
HISTORY_STAGE_FILE = HISTORY_DIR / "process_stage_daily.csv"
HISTORY_UNIT_FILE = HISTORY_DIR / "process_unit_daily.csv"
FALLBACK_SCOPE_FILE = REALTIME_DIR / "scope123_总汇总" / "latest_7d_daily.csv"
STAGE_ROOT = REALTIME_DIR / "process_stage_outputs"
FALLBACK_STAGE_FILE = STAGE_ROOT / "工艺段汇总" / "latest_7d_daily.csv"

BLUE = "123B75"
CYAN = "1CA6E8"
LIGHT_BLUE = "DCEEFF"
LIGHT_GRAY = "E9EEF5"
WHITE = "FFFFFF"
GRID = Side(style="thin", color="CCD6E3")


class ReportBody(BaseModel):
    startDate: date
    endDate: date


def _read_csv(path: Path, required_columns: list[str]) -> pd.DataFrame:
    if not path.exists():
        raise HTTPException(status_code=500, detail=f"未找到报告数据文件：{path}")

    try:
        df = pd.read_csv(path, encoding="utf-8-sig")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"报告数据读取失败：{path.name}，{exc}")

    missing = [column for column in required_columns if column not in df.columns]
    if missing:
        raise HTTPException(
            status_code=500,
            detail=f"报告数据文件 {path.name} 缺少字段：{missing}",
        )

    df = df.copy()
    df["period_start"] = pd.to_datetime(
        df["period_start"], errors="coerce", utc=True
    ).dt.tz_localize(None)
    return df.dropna(subset=["period_start"])


def _history_or_fallback(history_path: Path, fallback_path: Path) -> Path:
    """历史文件存在时优先使用，否则保持原来的滚动数据行为。"""
    return history_path if history_path.exists() else fallback_path


def _filter_dates(df: pd.DataFrame, start_date: date, end_date: date) -> pd.DataFrame:
    dates = df["period_start"].dt.date
    return df[(dates >= start_date) & (dates <= end_date)].copy()


def _load_report_data(
    start_date: date, end_date: date
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    scope_path = _history_or_fallback(HISTORY_SCOPE_FILE, FALLBACK_SCOPE_FILE)
    scope = _read_csv(
        scope_path,
        [
            "period_start",
            "scope1_carbon_kg",
            "scope2_carbon_kg",
            "scope3_carbon_kg",
            "total_carbon_kg",
        ],
    )
    scope_filtered = _filter_dates(scope, start_date, end_date)
    if scope_filtered.empty:
        available_start = scope["period_start"].min().date()
        available_end = scope["period_start"].max().date()
        raise HTTPException(
            status_code=404,
            detail=(
                f"所选日期没有可导出的碳排数据；当前可用日期为 "
                f"{available_start} 至 {available_end}"
            ),
        )

    stage = _read_csv(
        _history_or_fallback(HISTORY_STAGE_FILE, FALLBACK_STAGE_FILE),
        [
            "period_start",
            "process_stage",
            "stage_total_carbon_kg",
            "plant_total_carbon_kg",
            "stage_share_of_plant",
        ],
    )
    stage_filtered = _filter_dates(stage, start_date, end_date)

    if HISTORY_UNIT_FILE.exists():
        details = _read_csv(
            HISTORY_UNIT_FILE,
            [
                "period_start",
                "process_stage",
                "process_unit",
                "electric_carbon_kg",
                "chemical_carbon_kg",
                "unit_total_carbon_kg",
            ],
        )
        details = _filter_dates(details, start_date, end_date)
    else:
        detail_frames = []
        for path in sorted(STAGE_ROOT.glob("*/latest_7d_daily.csv")):
            if path.parent.name == "工艺段汇总":
                continue
            detail = _read_csv(
                path,
                [
                    "period_start",
                    "process_stage",
                    "process_unit",
                    "electric_carbon_kg",
                    "chemical_carbon_kg",
                    "unit_total_carbon_kg",
                ],
            )
            detail = _filter_dates(detail, start_date, end_date)
            if not detail.empty:
                detail_frames.append(detail)

        if detail_frames:
            details = pd.concat(detail_frames, ignore_index=True)
        else:
            details = pd.DataFrame(
                columns=[
                    "period_start",
                    "process_stage",
                    "process_unit",
                    "electric_carbon_kg",
                    "chemical_carbon_kg",
                    "unit_total_carbon_kg",
                ]
            )

    return (
        scope_filtered.sort_values("period_start"),
        stage_filtered.sort_values(["period_start", "process_stage"]),
        details.sort_values(["period_start", "process_stage", "process_unit"]),
    )


def _style_title(ws, title: str, end_column: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.font = Font(color=WHITE, bold=True, size=18)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32


def _style_header(ws, row: int, start_column: int, end_column: int) -> None:
    for column in range(start_column, end_column + 1):
        cell = ws.cell(row, column)
        cell.fill = PatternFill("solid", fgColor=CYAN)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=GRID, bottom=GRID, left=GRID, right=GRID)


def _style_table(ws, start_row: int, end_row: int, end_column: int) -> None:
    for row in ws.iter_rows(
        min_row=start_row, max_row=end_row, min_col=1, max_col=end_column
    ):
        for cell in row:
            cell.border = Border(top=GRID, bottom=GRID, left=GRID, right=GRID)
            cell.alignment = Alignment(vertical="center")


def _set_widths(ws, widths: dict[int, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = width


def _build_workbook(
    start_date: date,
    end_date: date,
    scope: pd.DataFrame,
    stage: pd.DataFrame,
    details: pd.DataFrame,
) -> BytesIO:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "报告汇总"
    daily_ws = wb.create_sheet("每日碳排")
    stage_ws = wb.create_sheet("工艺段汇总")
    detail_ws = wb.create_sheet("工艺单元明细")

    total_scope1 = float(scope["scope1_carbon_kg"].sum())
    total_scope2 = float(scope["scope2_carbon_kg"].sum())
    total_scope3 = float(scope["scope3_carbon_kg"].sum())
    total_carbon = float(scope["total_carbon_kg"].sum())
    daily_average = total_carbon / len(scope)
    peak_row = scope.loc[scope["total_carbon_kg"].idxmax()]

    _style_title(summary_ws, "黄阁水厂碳排放报告", 6)
    summary_ws["A3"] = "报告日期"
    summary_ws["B3"] = f"{start_date} 至 {end_date}"
    summary_ws["A4"] = "实际数据范围"
    summary_ws["B4"] = (
        f"{scope['period_start'].min().date()} 至 {scope['period_start'].max().date()}"
    )
    summary_ws["A5"] = "数据天数"
    summary_ws["B5"] = len(scope)
    for row in range(3, 6):
        summary_ws.cell(row, 1).font = Font(bold=True, color=BLUE)
        summary_ws.cell(row, 1).fill = PatternFill("solid", fgColor=LIGHT_BLUE)

    summary_ws.append([])
    summary_ws.append(
        ["指标", "范围1", "范围2", "范围3", "总碳排放", "日均碳排放"]
    )
    summary_ws.append(
        [
            "碳排放量（kgCO2e）",
            total_scope1,
            total_scope2,
            total_scope3,
            total_carbon,
            daily_average,
        ]
    )
    _style_header(summary_ws, 7, 1, 6)
    _style_table(summary_ws, 7, 8, 6)
    for cell in summary_ws[8][1:]:
        cell.number_format = '#,##0.00" kgCO2e"'

    summary_ws["A10"] = "最高日碳排放"
    summary_ws["B10"] = peak_row["period_start"].date()
    summary_ws["C10"] = float(peak_row["total_carbon_kg"])
    summary_ws["C10"].number_format = '#,##0.00" kgCO2e"'
    summary_ws["A10"].font = Font(bold=True, color=BLUE)

    stage_summary = (
        stage.groupby("process_stage", as_index=False)["stage_total_carbon_kg"].sum()
        if not stage.empty
        else pd.DataFrame(columns=["process_stage", "stage_total_carbon_kg"])
    )
    stage_summary["share"] = (
        stage_summary["stage_total_carbon_kg"] / total_carbon
        if total_carbon
        else 0
    )
    stage_start = 12
    summary_ws.cell(stage_start, 1, "工艺段")
    summary_ws.cell(stage_start, 2, "碳排放量（kgCO2e）")
    summary_ws.cell(stage_start, 3, "占全厂比例")
    for _, row in stage_summary.iterrows():
        summary_ws.append(
            [
                str(row["process_stage"]),
                float(row["stage_total_carbon_kg"]),
                float(row["share"]),
            ]
        )
    _style_header(summary_ws, stage_start, 1, 3)
    if len(stage_summary):
        _style_table(summary_ws, stage_start, stage_start + len(stage_summary), 3)
        for row in range(stage_start + 1, stage_start + len(stage_summary) + 1):
            summary_ws.cell(row, 2).number_format = '#,##0.00" kgCO2e"'
            summary_ws.cell(row, 3).number_format = "0.00%"
        chart = BarChart()
        chart.title = "各工艺段碳排放"
        chart.y_axis.title = "kgCO2e"
        chart.add_data(
            Reference(
                summary_ws,
                min_col=2,
                min_row=stage_start,
                max_row=stage_start + len(stage_summary),
            ),
            titles_from_data=True,
        )
        chart.set_categories(
            Reference(
                summary_ws,
                min_col=1,
                min_row=stage_start + 1,
                max_row=stage_start + len(stage_summary),
            )
        )
        chart.height = 8
        chart.width = 15
        summary_ws.add_chart(chart, "H3")

    _set_widths(summary_ws, {1: 28, 2: 22, 3: 18, 4: 18, 5: 20, 6: 20})
    summary_ws.freeze_panes = "A7"

    _style_title(daily_ws, "每日范围碳排放", 5)
    daily_ws.append(["日期", "范围1（kgCO2e）", "范围2（kgCO2e）", "范围3（kgCO2e）", "总碳排放（kgCO2e）"])
    for _, row in scope.iterrows():
        daily_ws.append(
            [
                row["period_start"].date(),
                float(row["scope1_carbon_kg"]),
                float(row["scope2_carbon_kg"]),
                float(row["scope3_carbon_kg"]),
                float(row["total_carbon_kg"]),
            ]
        )
    _style_header(daily_ws, 2, 1, 5)
    _style_table(daily_ws, 2, daily_ws.max_row, 5)
    for row in range(3, daily_ws.max_row + 1):
        daily_ws.cell(row, 1).number_format = "yyyy-mm-dd"
        for column in range(2, 6):
            daily_ws.cell(row, column).number_format = "#,##0.00"
    line_chart = LineChart()
    line_chart.title = "每日碳排放趋势"
    line_chart.y_axis.title = "kgCO2e"
    line_chart.x_axis.title = "日期"
    line_chart.add_data(
        Reference(daily_ws, min_col=2, max_col=5, min_row=2, max_row=daily_ws.max_row),
        titles_from_data=True,
    )
    line_chart.set_categories(
        Reference(daily_ws, min_col=1, min_row=3, max_row=daily_ws.max_row)
    )
    line_chart.height = 8
    line_chart.width = 17
    daily_ws.add_chart(line_chart, "G2")
    _set_widths(daily_ws, {1: 16, 2: 20, 3: 20, 4: 20, 5: 22})
    daily_ws.freeze_panes = "A3"
    daily_ws.auto_filter.ref = f"A2:E{daily_ws.max_row}"

    _style_title(stage_ws, "工艺段每日碳排放", 5)
    stage_ws.append(
        ["日期", "工艺段", "工艺段碳排放（kgCO2e）", "全厂碳排放（kgCO2e）", "工艺段占比"]
    )
    for _, row in stage.iterrows():
        stage_ws.append(
            [
                row["period_start"].date(),
                str(row["process_stage"]),
                float(row["stage_total_carbon_kg"]),
                float(row["plant_total_carbon_kg"]),
                float(row["stage_share_of_plant"]),
            ]
        )
    _style_header(stage_ws, 2, 1, 5)
    _style_table(stage_ws, 2, stage_ws.max_row, 5)
    for row in range(3, stage_ws.max_row + 1):
        stage_ws.cell(row, 1).number_format = "yyyy-mm-dd"
        stage_ws.cell(row, 3).number_format = "#,##0.00"
        stage_ws.cell(row, 4).number_format = "#,##0.00"
        stage_ws.cell(row, 5).number_format = "0.00%"
    _set_widths(stage_ws, {1: 16, 2: 25, 3: 25, 4: 24, 5: 18})
    stage_ws.freeze_panes = "A3"
    stage_ws.auto_filter.ref = f"A2:E{stage_ws.max_row}"

    _style_title(detail_ws, "工艺单元每日碳排放明细", 6)
    detail_ws.append(
        [
            "日期",
            "工艺段",
            "工艺单元",
            "电耗碳排（kgCO2e）",
            "药耗碳排（kgCO2e）",
            "单元总碳排（kgCO2e）",
        ]
    )
    for _, row in details.iterrows():
        detail_ws.append(
            [
                row["period_start"].date(),
                str(row["process_stage"]),
                str(row["process_unit"]),
                float(row["electric_carbon_kg"]),
                float(row["chemical_carbon_kg"]),
                float(row["unit_total_carbon_kg"]),
            ]
        )
    _style_header(detail_ws, 2, 1, 6)
    _style_table(detail_ws, 2, detail_ws.max_row, 6)
    for row in range(3, detail_ws.max_row + 1):
        detail_ws.cell(row, 1).number_format = "yyyy-mm-dd"
        for column in range(4, 7):
            detail_ws.cell(row, column).number_format = "#,##0.00"
    _set_widths(detail_ws, {1: 16, 2: 25, 3: 42, 4: 23, 5: 23, 6: 24})
    detail_ws.freeze_panes = "A3"
    detail_ws.auto_filter.ref = f"A2:F{detail_ws.max_row}"

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.post("/api/report/export", summary="导出碳排放报告")
def export_report(body: ReportBody):
    if body.endDate < body.startDate:
        raise HTTPException(status_code=400, detail="endDate 不能早于 startDate")
    if (body.endDate - body.startDate).days > 366:
        raise HTTPException(status_code=400, detail="单次报告日期范围不能超过 366 天")

    scope, stage, details = _load_report_data(body.startDate, body.endDate)
    output = _build_workbook(body.startDate, body.endDate, scope, stage, details)
    filename = f"黄阁水厂碳排放报告_{body.startDate}_{body.endDate}.xlsx"
    disposition = f"attachment; filename*=UTF-8''{quote(filename)}"

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": disposition},
    )
